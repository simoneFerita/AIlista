#!/usr/bin/env python3
"""
Ailista Ultra – Table Only PRO Edition (Oggetto Column)
OCR + Estrazione campi + Tabella ordinata + CSV/JSON/XLSX
(Senza colonna File, Reso vuota, Preview sostituita con Oggetto)
"""

import sys
import re
import json
import csv
import numpy as np
from pathlib import Path
from typing import List, Dict

from PIL import Image, ImageFilter, ImageOps
import easyocr

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors


class AilistaTablePRO:
    def __init__(self, input_dir: str, output_pdf: str):
        self.input_dir = Path(input_dir)
        self.output_pdf = Path(output_pdf)
        self.reader = easyocr.Reader(["it", "en"], gpu=False)
        self.items: List[Dict] = []

    def preprocess(self, img: Image.Image) -> Image.Image:
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = img.filter(ImageFilter.SHARPEN)
        w, h = img.size
        if w < 1000:
            img = img.resize((w * 2, h * 2), Image.LANCZOS)
        return img

    def extract_text(self, img: Image.Image) -> str:
        img = self.preprocess(img)
        img_np = np.array(img)
        results = self.reader.readtext(img_np, detail=1, paragraph=False)
        blocks = [(bbox[0][1], bbox[0][0], text) for (bbox, text, conf) in results]
        blocks.sort(key=lambda b: (b[0], b[1]))
        return "\n".join([b[2] for b in blocks]).strip()

    def extract_price(self, text: str) -> float:
        m = re.search(r"(\d{1,4}[.,]\d{2})", text)
        return float(m.group(1).replace(",", ".")) if m else 0.0

    def extract_vendor(self, text: str) -> str:
        m = re.search(r"Venduto da:\s*([A-Za-z0-9 ]]+)", text)
        if m:
            return m.group(1).strip()
        if "Amazon" in text:
            return "Amazon"
        if "Chen" in text:
            return "ChenYang Cable"
        if "Jimi" in text:
            return "Jimi Cable"
        return "N/D"

    def extract_condition(self, text: str) -> str:
        m = re.search(r"Condizione:\s*([A-Za-z0-9: ]]+)", text)
        if m:
            return m.group(1).strip()
        if "ottime" in text:
            return "Usato: ottime condizioni"
        if "buone" in text:
            return "Usato: buone condizioni"
        if "accettabile" in text:
            return "Usato: accettabile"
        if "nuovo" in text:
            return "Usato: nuovo di zecca"
        return "N/D"

    def extract_return(self, text: str) -> str:
        return ""  # richiesto

    def extract_object(self, text: str) -> str:
        first_line = text.split("\n")[0]
        first_line = re.sub(r"[^A-Za-z0-9 ]", "", first_line)
        return first_line.strip()

    def analyze_image(self, img_path: Path):
        img = Image.open(img_path)
        text = self.extract_text(img)
        self.items.append({
            "object": self.extract_object(text),
            "text": text,
            "price": self.extract_price(text),
            "vendor": self.extract_vendor(text),
            "condition": self.extract_condition(text),
            "returnable": self.extract_return(text)
        })

    def export_csv(self):
        csv_path = self.output_pdf.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Oggetto", "Prezzo", "Venditore", "Condizione", "Reso"])
            for item in self.items:
                writer.writerow([
                    item["object"],
                    f"{item['price']:.2f}",
                    item["vendor"],
                    item["condition"],
                    item["returnable"]
                ])

    def export_json(self):
        json_path = self.output_pdf.with_suffix(".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(self.items, f, indent=4, ensure_ascii=False)

    def export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, PieChart

        xlsx_path = self.output_pdf.with_suffix(".xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Oggetti"

        headers = ["Oggetto", "Prezzo", "Venditore", "Condizione", "Reso"]
        ws.append(headers)

        # Header style
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for item in self.items:
            ws.append([
                item["object"],
                item["price"],
                item["vendor"],
                item["condition"],
                item["returnable"]
            ])

        # Cell formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Filters
        ws.auto_filter.ref = f"A1:E{ws.max_row}"

        # Sort by price (B column)
        ws.auto_filter.add_sort_condition(f"B2:B{ws.max_row}")

        # Bar chart: prices
        chart = BarChart()
        chart.title = "Prezzi degli Oggetti"
        chart.y_axis.title = "Prezzo (€)"
        chart.x_axis.title = "Oggetto"

        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "G2")

        # Pie chart: vendors
        vendor_sheet = wb.create_sheet("Venditori")
        vendor_sheet.append(["Venditore", "Conteggio"])

        vendor_count = {}
        for item in self.items:
            vendor_count[item["vendor"]] = vendor_count.get(item["vendor"], 0) + 1

        for vendor, count in vendor_count.items():
            vendor_sheet.append([vendor, count])

        pie = PieChart()
        pie.title = "Distribuzione Venditori"

        data = Reference(vendor_sheet, min_col=2, min_row=1, max_row=len(vendor_count) + 1)
        labels = Reference(vendor_sheet, min_col=1, min_row=2, max_row=len(vendor_count) + 1)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 10
        pie.width = 10

        vendor_sheet.add_chart(pie, "D2")

        wb.save(xlsx_path)

    def generate_pdf(self):
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(self.output_pdf), pagesize=A4,
                                rightMargin=40, leftMargin=40,
                                topMargin=40, bottomMargin=40)

        elements = []
        title = Paragraph("Lista Oggetti Ordinata per Prezzo", styles["Title"])
        title.alignment = TA_CENTER
        elements.append(title)
        elements.append(Spacer(1, 20))

        self.items.sort(key=lambda x: x["price"], reverse=True)

        table_data = [["#", "Oggetto", "Prezzo (€)", "Venditore", "Condizione", "Reso"]]

        for i, item in enumerate(self.items, start=1):
            table_data.append([
                str(i),
                item["object"],
                f"{item['price']:.2f}",
                item["vendor"],
                item["condition"],
                item["returnable"]
            ])

        table = Table(table_data, colWidths=[30, 200, 70, 100, 140, 40])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))

        elements.append(table)
        doc.build(elements)
        return str(self.output_pdf)

    def process(self):
        images = [f for f in self.input_dir.iterdir()
                  if f.suffix.lower() in [".jpg", ".jpeg", ".png", ".bmp"]]

        for img_path in images:
            self.analyze_image(img_path)

        pdf_path = self.generate_pdf()
        self.export_csv()
        self.export_json()
        self.export_excel()
        return pdf_path


if __name__ == "__main__":
    input_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    output_pdf = sys.argv[2] if len(sys.argv) > 2 else "output.pdf"
    app = AilistaTablePRO(input_dir, output_pdf)
    app.process()
